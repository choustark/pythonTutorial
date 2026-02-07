"""
使用 openpyxl 库 操作excel文件，
openpyxl并不支持操作 Office 2007 以前版本的 Excel 文件。
"""

import openpyxl, datetime


def create_excel_file(file_name: str, sheet_name: str,
                      header: list[str], data: list[list[str]]) -> None:
    """
    创建excel文件
    :param file_name:
    :param sheet_name:
    :param header
    :param data
    :return:
    """
    if file_name is None or sheet_name is None:
        raise ValueError("参数不能为空")
    if not file_name.endswith('.xlsx'):
        raise ValueError("文件格式错误")
    if header is None or data is None:
        raise ValueError("表头和数据不能为空")
    # 创建工作薄
    wb: openpyxl.Workbook = openpyxl.Workbook()
    # 添加工作表
    sheet = wb.active
    sheet.title = sheet_name
    for index, head in enumerate(header):
        # 写入表头,参数为行，列，值
        sheet.cell(row=1, column=index + 1, value=head).alignment = openpyxl.styles.Alignment(horizontal='center',
                                                                                              vertical='center')

    for data_index, item in enumerate(data):
        # 填充数据
        for index, datum in enumerate(item):
            sheet.cell(row=data_index + 2, column=index + 1, value=datum).alignment = openpyxl.styles.Alignment(
                horizontal='center',
                vertical='center')
    # 保存文件
    wb.save(file_name)


# 读取excel文件
def read_excel_file(file_name: str) -> list[list[str]]:
    """
    读取excel 文件
    :param file_name: 文件名称
    :return: 读取到的数据
    """
    if file_name is None:
        raise ValueError("参数不能为空")
    if not file_name.endswith('.xlsx'):
        raise ValueError("文件格式错误")
    wb = openpyxl.load_workbook(file_name)
    # 获取工作表的名字
    sheet_names = wb.sheetnames
    print(sheet_names)
    # 获取工作表 ---> Worksheet
    sheet = wb.worksheets[0]
    print(sheet.dimensions)
    print(sheet.max_row, sheet.max_column)
    # 获取指定单元格
    print(sheet.cell(row=1, column=1).value)
    print(sheet['C3'].value)
    # 获取多个单元格
    print(sheet['A1':'C3'])
    file_data: list[list[str]] = []

    for row in range(2, sheet.max_row + 1):
        row_data: list[str] = []
        for col_ch in 'ABCDE':
            value = sheet[f'{col_ch}{row}'].value
            if type(value) == datetime.datetime:
                print(value.strftime('%Y年%m月%d日'), end='\t')
            elif type(value) == int:
                print(f'{value:<10d}', end='\t')
            elif type(value) == float:
                print(f'{value:.4f}', end='\t')
            else:
                print(value, end='\t')
            row_data.append(value)
        print()
        file_data.append(row_data)
        # row_data.clear()

    return file_data


header = ['序号', '姓名', '年龄', '性别', "描述"]
data = [
    [1, '关羽', 18, '男', '三国时期蜀汉名将，以忠义著称'],
    [2, '赵云', 19, '女', '三国时期蜀汉将领，勇猛善战'],
    [3, '马超', 20, '男', '三国时期蜀汉将领，西凉猛将'],
    [4, '黄忠', 18, '男', '三国时期蜀汉老将，箭术精湛'],
    [5, '刘备', 19, '女', '三国时期蜀汉开国皇帝'],
    [6, '夏侯惇', 20, '男', '三国时期魏国名将，曹操族弟'],
    [7, '吕布', 18, '男', '三国时期著名武将，武艺高强'],
    [8, '诸葛亮', 19, '女', '三国时期蜀汉丞相，智慧化身'],
    [8, '貂蝉', 26, '女', '三国时期美女，连环计主角'],
]

# create_excel_file('三国花名册.xlsx', '三国英雄', header, data)

print(read_excel_file('三国花名册.xlsx'))
